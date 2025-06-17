from openpyxl import load_workbook
import os

# テンプレートファイルの数式を確認
current_dir = os.path.dirname(os.path.abspath(__file__))
template_path = os.path.join(current_dir, '生成するファイルの例', '2025_04_高速道路等利用実績簿（テンプレート）.xlsx')

wb = load_workbook(template_path, data_only=False)
ws = wb.active

print("=== テンプレートの重要なセルの内容と数式 ===")

# 年月関連のセル
important_cells = ['B5', 'D5', 'E56', 'E57', 'B13', 'C13', 'J13', 'K13', 'M7']

for cell_ref in important_cells:
    cell = ws[cell_ref]
    print(f"{cell_ref}:")
    print(f"  値: {cell.value}")
    print(f"  数式: {cell.formula if hasattr(cell, 'formula') else 'なし'}")
    print(f"  データ型: {type(cell.value)}")
    print()

# 日付計算に使われているセル
print("=== 日付計算セル（B13-B20）===")
for row in range(13, 21):
    cell = ws[f'B{row}']
    print(f"B{row}: {cell.value}")

print("\n=== 曜日計算セル（C13-C20）===")
for row in range(13, 21):
    cell = ws[f'C{row}']
    print(f"C{row}: {cell.value}")

# E56, E57の値を確認
print(f"\n=== 基準日付 ===")
print(f"E56 (開始日): {ws['E56'].value}")
print(f"E57 (終了日): {ws['E57'].value}")

# M7の数式も確認
print(f"\n=== M7の数式 ===")
print(f"M7: {ws['M7'].value}")