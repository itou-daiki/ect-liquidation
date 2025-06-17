import streamlit as st
import pandas as pd
import chardet
from datetime import datetime
import io
import requests
import json
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import calendar
import shutil

# アプリケーションの設定
st.set_page_config(
    page_title="高速道路利用実績簿生成",
    page_icon="🛣️",
    layout="wide"
)

def detect_encoding(uploaded_file):
    """アップロードされたファイルのエンコーディングを検出"""
    raw_data = uploaded_file.read()
    uploaded_file.seek(0)  # ファイルポインタをリセット
    encoding = chardet.detect(raw_data)
    return encoding['encoding']

def load_csv_data(uploaded_file, encoding):
    """CSVファイルを読み込む"""
    try:
        df = pd.read_csv(uploaded_file, encoding=encoding)
        return df
    except Exception as e:
        st.error(f"CSVファイルの読み込みに失敗しました: {e}")
        return None

def extract_year_month(df):
    """データから年月を抽出（最新の年月を優先）"""
    if '利用年月日（自）' in df.columns:
        year_months = []
        
        # 全ての日付を確認して年月を収集
        for date_str in df['利用年月日（自）'].dropna():
            if '/' in str(date_str):
                parts = str(date_str).split('/')
                if len(parts) >= 2:
                    try:
                        year = int(parts[0])
                        month = int(parts[1])
                        # 2桁年を4桁年に変換
                        if year < 50:  # 25年以下は2025年以降と仮定
                            year += 2000
                        elif year < 100:  # 50-99年は1950-1999年と仮定
                            year += 1900
                        year_months.append((year, month))
                    except ValueError:
                        continue
        
        if year_months:
            # 最新の年月を返す
            year_months.sort(reverse=True)
            return year_months[0]
    
    return None, None

def get_highway_sections():
    """高速道路区間のリストを取得"""
    # 九州地方の主要IC・SA・PA
    sections = [
        "大分米良",
        "日田",
        "福岡",
        "北九州",
        "熊本",
        "鹿児島",
        "宮崎",
        "佐賀",
        "長崎",
        "別府",
        "大分",
        "中津",
        "玖珠",
        "天瀬高塚",
        "杷木",
        "筑紫野",
        "太宰府",
        "春日",
        "古賀",
        "宗像",
        "若宮",
        "飯塚",
        "八幡",
        "小倉",
        "門司",
        "下関",
        "美祢",
        "山口",
        "防府",
        "徳山",
        "岩国"
    ]
    return sorted(sections)

def match_csv_to_template(df, year, month, highway_from, highway_to):
    """CSVデータをテンプレートの入力可能箇所にマッチング"""
    from datetime import datetime, time
    import calendar
    
    # 月の日数を取得
    last_day = calendar.monthrange(year, month)[1]
    
    # 入力データの初期化
    template_data = {
        'header_info': {
            'organization': '',  # C3
            'position': '',      # K3  
            'name': ''           # N3
        },
        'date_month': {
            'year': year - 2018,  # B5 (令和年)
            'month': month        # D5
        },
        'highway_info': {
            'from': highway_from,  # M5
            'to': highway_to,      # P5
            'one_way_fee': 2680    # M6
        },
        'daily_data': {}  # 日別の利用データ
    }
    
    # 日別データの初期化
    for day in range(1, last_day + 1):
        template_data['daily_data'][day] = {
            'outbound_confirmed': None,  # D列（往路利用確認）
            'outbound_amount': 0,        # E列（往路利用金額）
            'return_confirmed': None,    # G列（復路利用確認）
            'return_amount': 0,          # H列（復路利用金額）
            'morning_trips': [],         # 午前中の利用記録
            'afternoon_trips': []        # 午後の利用記録
        }
    
    # CSVデータを解析して日別データにマッピング
    for index, row in df.iterrows():
        date_str = row['利用年月日（自）']
        time_str = row['時分（自）']
        
        try:
            if '/' in date_str:
                parts = date_str.split('/')
                if len(parts) >= 3:
                    year_part = int(parts[0])
                    month_part = int(parts[1])
                    day_part = int(parts[2])
                    
                    # 年を正規化
                    if year_part < 50:
                        year_part += 2000
                    elif year_part < 100:
                        year_part += 1900
                    
                    # 指定された年月と一致するかチェック
                    if year_part == year and month_part == month and 1 <= day_part <= last_day:
                        day = day_part
                        
                        # 時刻の解析（HH:MM形式）
                        is_morning = True  # デフォルトは午前
                        if ':' in str(time_str):
                            try:
                                hour = int(str(time_str).split(':')[0])
                                is_morning = hour < 12  # 12時未満は午前
                            except:
                                is_morning = True
                        
                        # 往復判定と利用区間チェック
                        from_ic = str(row['利用ＩＣ（自）'])
                        to_ic = str(row['利用ＩＣ（至）'])
                        amount = row['通行料金']
                        
                        # 利用区間内での利用かチェック
                        is_in_route = False
                        is_traffic_stop = False
                        
                        if highway_from in from_ic and highway_to in to_ic:
                            # 往路（出発地→到着地）
                            is_in_route = True
                            # 利用区間内での下車チェック（同じIC間での利用は通行止の可能性）
                            if from_ic == to_ic or (highway_from in to_ic and highway_to in from_ic):
                                is_traffic_stop = True
                            
                            if is_morning:
                                template_data['daily_data'][day]['morning_trips'].append(amount)
                                template_data['daily_data'][day]['outbound_confirmed'] = '通行止' if is_traffic_stop else '○'
                            else:
                                template_data['daily_data'][day]['afternoon_trips'].append(amount)
                                template_data['daily_data'][day]['return_confirmed'] = '通行止' if is_traffic_stop else '○'
                                
                        elif highway_to in from_ic and highway_from in to_ic:
                            # 復路（到着地→出発地）
                            is_in_route = True
                            # 利用区間内での下車チェック
                            if from_ic == to_ic or (highway_to in to_ic and highway_from in from_ic):
                                is_traffic_stop = True
                            
                            if is_morning:
                                template_data['daily_data'][day]['morning_trips'].append(amount)
                                template_data['daily_data'][day]['outbound_confirmed'] = '通行止' if is_traffic_stop else '○'
                            else:
                                template_data['daily_data'][day]['afternoon_trips'].append(amount)
                                template_data['daily_data'][day]['return_confirmed'] = '通行止' if is_traffic_stop else '○'
                        
                        # 利用区間内での利用があった場合の料金集計
                        elif highway_from in from_ic or highway_to in from_ic or highway_from in to_ic or highway_to in to_ic:
                            # 利用区間に関連する利用
                            if is_morning:
                                template_data['daily_data'][day]['morning_trips'].append(amount)
                                if not template_data['daily_data'][day]['outbound_confirmed']:
                                    template_data['daily_data'][day]['outbound_confirmed'] = '○'
                            else:
                                template_data['daily_data'][day]['afternoon_trips'].append(amount)
                                if not template_data['daily_data'][day]['return_confirmed']:
                                    template_data['daily_data'][day]['return_confirmed'] = '○'
                                    
        except (ValueError, IndexError):
            continue
    
    # 日別料金の合計を計算
    for day in range(1, last_day + 1):
        day_data = template_data['daily_data'][day]
        
        # 午前中の利用額合計（往路）
        if day_data['morning_trips']:
            day_data['outbound_amount'] = sum(day_data['morning_trips'])
        
        # 午後の利用額合計（復路）
        if day_data['afternoon_trips']:
            day_data['return_amount'] = sum(day_data['afternoon_trips'])
    
    return template_data

def calculate_usage_amount_by_date_formula(df, year, month):
    """=DATE(year,month,day)アルゴリズムで日付とマッチしている箇所の利用金額を算出"""
    from datetime import datetime
    import calendar
    
    # 月の日数を取得
    last_day = calendar.monthrange(year, month)[1]
    
    # 処理開始
    print(f"Processing {df.shape[0]} records for {year}-{month:02d}")
    
    # 日付マッチングによる利用金額の計算
    usage_amounts = {}
    
    # 1日から月末まで全ての日付をチェック
    for day in range(1, last_day + 1):
        target_date = datetime(year, month, day)
        usage_amounts[day] = calculate_daily_usage_from_csv(df, target_date)
    
    return usage_amounts

def calculate_daily_usage_from_csv(df, target_date, one_way_fee=2680):
    """指定された日付のCSVデータから利用金額を計算"""
    morning_amount = 0
    afternoon_amount = 0
    morning_confirmed = None
    afternoon_confirmed = None
    
    # 複数の日付フォーマットに対応
    target_date_formats = []
    
    # 基本フォーマット
    target_date_formats.append(target_date.strftime('%y/%m/%d'))    # 25/05/01
    target_date_formats.append(target_date.strftime('%Y/%m/%d'))    # 2025/05/01
    
    # ゼロパディングなしフォーマット（プラットフォーム依存を回避）
    y2 = target_date.strftime('%y')
    y4 = target_date.strftime('%Y')
    m_padded = target_date.strftime('%m')
    m_no_pad = str(target_date.month)
    d_padded = target_date.strftime('%d')
    d_no_pad = str(target_date.day)
    
    target_date_formats.extend([
        f"{y2}/{m_no_pad}/{d_no_pad}",      # 25/5/1
        f"{y4}/{m_no_pad}/{d_no_pad}",      # 2025/5/1
        f"{y2}/{m_padded}/{d_no_pad}",      # 25/05/1
        f"{y4}/{m_padded}/{d_no_pad}",      # 2025/05/1
        f"{y2}/{m_no_pad}/{d_padded}",      # 25/5/01
        f"{y4}/{m_no_pad}/{d_padded}",      # 2025/5/01
    ])
    
    # 日付フォーマット確認（初回のみ）
    if target_date.day == 1:
        print(f"Matching date format example: {target_date_formats[0]}")
    
    for _, row in df.iterrows():
        # 直接列名でアクセス（CSVの構造が確認済みのため）
        try:
            date_str = str(row['利用年月日（自）']).strip()
            time_str = str(row['時分（自）']).strip()
            amount = float(row['通行料金'])
        except (KeyError, ValueError, TypeError):
            # フォールバック：インデックスベースで取得
            try:
                date_str = str(row.iloc[0]).strip()
                time_str = str(row.iloc[1]).strip()
                amount = float(row.iloc[8])
            except (ValueError, TypeError, IndexError):
                continue
        
        # 日付の正規化とマッチング（複数フォーマット対応）
        date_match = False
        for target_format in target_date_formats:
            if date_str == target_format:
                date_match = True
                break
        
        # 手動での日付パース（より柔軟な対応）
        if not date_match and '/' in date_str:
            try:
                parts = date_str.split('/')
                if len(parts) >= 3:
                    csv_year = int(parts[0])
                    csv_month = int(parts[1])
                    csv_day = int(parts[2])
                    
                    # 年の正規化
                    if csv_year < 50:
                        csv_year += 2000
                    elif csv_year < 100:
                        csv_year += 1900
                    
                    if (csv_year == target_date.year and 
                        csv_month == target_date.month and 
                        csv_day == target_date.day):
                        date_match = True
            except (ValueError, IndexError):
                continue
        
        if date_match and amount > 0:
            
            # 時刻による午前/午後の判定
            is_morning = True
            if ':' in time_str:
                try:
                    hour = int(time_str.split(':')[0])
                    is_morning = hour < 12
                except:
                    is_morning = True
            
            if is_morning:
                morning_amount += amount
                morning_confirmed = '○'
            else:
                afternoon_amount += amount
                afternoon_confirmed = '○'
    
    result = {
        'morning_amount': morning_amount,
        'afternoon_amount': afternoon_amount,
        'morning_confirmed': morning_confirmed,
        'afternoon_confirmed': afternoon_confirmed
    }
    
    # 処理結果（データがある場合のみ）
    if morning_amount > 0 or afternoon_amount > 0:
        print(f"✓ {target_date.strftime('%m/%d')}: Morning={morning_amount}, Afternoon={afternoon_amount}")
    
    return result

def generate_expense_report_from_template(df, year, month, highway_from, highway_to, one_way_fee, monthly_allowance, organization="", position="", name=""):
    """テンプレートファイルをベースに利用実績簿を生成（水色箇所のみ入力）"""
    
    # テンプレートファイルをコピー
    import os
    from datetime import datetime
    current_dir = os.path.dirname(os.path.abspath(__file__))
    template_path = os.path.join(current_dir, '生成するファイルの例', '2025_04_高速道路等利用実績簿（テンプレート）.xlsx')
    
    if not os.path.exists(template_path):
        raise FileNotFoundError(f"テンプレートファイルが見つかりません: {template_path}")
    
    wb = load_workbook(template_path, data_only=False)  # 数式を保持
    ws = wb.active
    
    # ヘッダー情報
    if organization:
        ws['C3'] = organization
    if position:
        ws['K3'] = position  
    if name:
        ws['N3'] = name
    
    # 日付情報を設定（E56, E57に月の初日と最終日を設定）
    ws['B5'] = year - 2018  # 令和年
    ws['D5'] = month
    
    # 計算用の日付設定
    first_day = datetime(year, month, 1)
    last_day_num = calendar.monthrange(year, month)[1]
    last_day = datetime(year, month, last_day_num)
    ws['E56'] = first_day
    ws['E57'] = last_day
    
    # 高速道路情報
    ws['M5'] = highway_from
    ws['P5'] = highway_to
    ws['M6'] = one_way_fee
    
    # DATE(year,month,day)アルゴリズムで利用金額を算出
    usage_amounts = calculate_usage_amount_by_date_formula(df, year, month)
    
    # 利用金額をExcelテンプレートに転記
    transferred_count = 0
    
    for day in range(1, last_day_num + 1):
        if day in usage_amounts:
            day_data = usage_amounts[day]
            
            # データがある場合のみ処理
            if (day_data['morning_amount'] > 0 or day_data['afternoon_amount'] > 0 or 
                day_data['morning_confirmed'] or day_data['afternoon_confirmed']):
                
                transferred_count += 1
                
                # 前半15日（13-27行）- 左側
                if day <= 15:
                    row = day + 12  # 1日→13行, 2日→14行, ...
                    
                    # 往路データ
                    if day_data['morning_confirmed']:
                        ws[f'D{row}'] = day_data['morning_confirmed']
                    if day_data['morning_amount'] > 0:
                        ws[f'E{row}'] = day_data['morning_amount']
                    
                    # 復路データ
                    if day_data['afternoon_confirmed']:
                        ws[f'G{row}'] = day_data['afternoon_confirmed']
                    if day_data['afternoon_amount'] > 0:
                        ws[f'H{row}'] = day_data['afternoon_amount']
                
                # 後半（16-31日）- 右側
                elif day >= 16:
                    if day <= 30:
                        row = day - 15 + 12  # 16日→13行, 17日→14行, ..., 30日→27行
                    else:  # 31日
                        row = 28
                        
                    # 往路データ（右側）
                    if day_data['morning_confirmed']:
                        ws[f'L{row}'] = day_data['morning_confirmed']
                    if day_data['morning_amount'] > 0:
                        ws[f'M{row}'] = day_data['morning_amount']
                    
                    # 復路データ（右側）
                    if day_data['afternoon_confirmed']:
                        ws[f'O{row}'] = day_data['afternoon_confirmed']
                    if day_data['afternoon_amount'] > 0:
                        ws[f'P{row}'] = day_data['afternoon_amount']
    
    print(f"✓ Transferred {transferred_count} days of usage data to Excel template")
    
    return wb

def generate_expense_report_from_template_legacy(df, year, month, highway_from, highway_to, one_way_fee, monthly_allowance, organization="", position="", name=""):
    """従来の方式によるテンプレートファイル生成（レガシー版）"""
    
    # テンプレートファイルをコピー
    import os
    from datetime import datetime
    current_dir = os.path.dirname(os.path.abspath(__file__))
    template_path = os.path.join(current_dir, '生成するファイルの例', '2025_04_高速道路等利用実績簿（テンプレート）.xlsx')
    
    if not os.path.exists(template_path):
        raise FileNotFoundError(f"テンプレートファイルが見つかりません: {template_path}")
    
    wb = load_workbook(template_path, data_only=False)  # 数式を保持
    ws = wb.active
    
    # ヘッダー情報
    if organization:
        ws['C3'] = organization
    if position:
        ws['K3'] = position  
    if name:
        ws['N3'] = name
    
    # 日付情報を設定（E56, E57に月の初日と最終日を設定）
    ws['B5'] = year - 2018  # 令和年
    ws['D5'] = month
    
    # 計算用の日付設定
    first_day = datetime(year, month, 1)
    last_day_num = calendar.monthrange(year, month)[1]
    last_day = datetime(year, month, last_day_num)
    ws['E56'] = first_day
    ws['E57'] = last_day
    
    # 高速道路情報
    ws['M5'] = highway_from
    ws['P5'] = highway_to
    ws['M6'] = one_way_fee
    
    # CSVデータを日付ごとに整理
    daily_usage = {}
    
    for index, row in df.iterrows():
        date_str = row['利用年月日（自）']
        time_str = row['時分（自）']
        
        try:
            if '/' in date_str:
                parts = date_str.split('/')
                if len(parts) >= 3:
                    year_part = int(parts[0])
                    month_part = int(parts[1])
                    day_part = int(parts[2])
                    
                    # 年を正規化
                    if year_part < 50:
                        year_part += 2000
                    elif year_part < 100:
                        year_part += 1900
                    
                    # 指定された年月と一致するかチェック
                    if year_part == year and month_part == month:
                        day = day_part
                        
                        if day not in daily_usage:
                            daily_usage[day] = {
                                'morning_trips': [],
                                'afternoon_trips': [],
                                'outbound_confirmed': None,
                                'return_confirmed': None
                            }
                        
                        # 時刻の解析（HH:MM形式）
                        is_morning = True
                        if ':' in str(time_str):
                            try:
                                hour = int(str(time_str).split(':')[0])
                                is_morning = hour < 12
                            except:
                                is_morning = True
                        
                        # 往復判定と利用区間チェック
                        from_ic = str(row['利用ＩＣ（自）'])
                        to_ic = str(row['利用ＩＣ（至）'])
                        amount = row['通行料金']
                        
                        # 利用区間との一致チェック
                        is_route_match = False
                        is_traffic_stop = False
                        
                        if highway_from in from_ic and highway_to in to_ic:
                            # 往路
                            is_route_match = True
                            if from_ic == to_ic:
                                is_traffic_stop = True
                        elif highway_to in from_ic and highway_from in to_ic:
                            # 復路
                            is_route_match = True
                            if from_ic == to_ic:
                                is_traffic_stop = True
                        elif highway_from in from_ic or highway_to in from_ic or highway_from in to_ic or highway_to in to_ic:
                            # 利用区間に関連
                            is_route_match = True
                        
                        if is_route_match:
                            if is_morning:
                                daily_usage[day]['morning_trips'].append(amount)
                                daily_usage[day]['outbound_confirmed'] = '通行止' if is_traffic_stop else '○'
                            else:
                                daily_usage[day]['afternoon_trips'].append(amount)
                                daily_usage[day]['return_confirmed'] = '通行止' if is_traffic_stop else '○'
                                
        except (ValueError, IndexError):
            continue
    
    # テンプレートの日付セルに基づいてデータを入力
    # 前半15日（13-27行）- 左側
    for row in range(13, 28):
        day = row - 12  # 1日から15日
        
        if day in daily_usage:
            data = daily_usage[day]
            
            # 往路データ
            if data['outbound_confirmed']:
                ws[f'D{row}'] = data['outbound_confirmed']
            if data['morning_trips']:
                ws[f'E{row}'] = sum(data['morning_trips'])
            
            # 復路データ
            if data['return_confirmed']:
                ws[f'G{row}'] = data['return_confirmed']
            if data['afternoon_trips']:
                ws[f'H{row}'] = sum(data['afternoon_trips'])
    
    # 後半（16-31日）- 右側
    for row in range(13, 28):
        day = (row - 12) + 15  # 16日から30日
        
        if day <= last_day_num and day in daily_usage:
            data = daily_usage[day]
            
            # 往路データ（右側）
            if data['outbound_confirmed']:
                ws[f'L{row}'] = data['outbound_confirmed']
            if data['morning_trips']:
                ws[f'M{row}'] = sum(data['morning_trips'])
            
            # 復路データ（右側）
            if data['return_confirmed']:
                ws[f'O{row}'] = data['return_confirmed']
            if data['afternoon_trips']:
                ws[f'P{row}'] = sum(data['afternoon_trips'])
    
    # 28行目の右側（31日）の処理
    if 31 <= last_day_num and 31 in daily_usage:
        data = daily_usage[31]
        
        if data['outbound_confirmed']:
            ws['L28'] = data['outbound_confirmed']
        if data['morning_trips']:
            ws['M28'] = sum(data['morning_trips'])
        
        if data['return_confirmed']:
            ws['O28'] = data['return_confirmed']
        if data['afternoon_trips']:
            ws['P28'] = sum(data['afternoon_trips'])
    
    return wb

def main():
    st.title("🛣️ 高速道路利用実績簿生成システム")
    st.markdown("---")
    
    # サイドバーで設定
    st.sidebar.header("設定")
    
    # 基本情報設定
    st.sidebar.header("基本情報")
    organization = st.sidebar.text_input("所属", value="日田高等学校")
    position = st.sidebar.text_input("職", value="教諭")
    name = st.sidebar.text_input("氏名", value="伊藤大貴")
    
    st.sidebar.header("利用区間設定")
    # 高速道路区間選択
    highway_sections = get_highway_sections()
    
    # 「大分米良」と「日田」のインデックスを取得
    try:
        oita_index = highway_sections.index("大分米良")
    except ValueError:
        oita_index = 0
    
    try:
        hita_index = highway_sections.index("日田")
    except ValueError:
        hita_index = 1
    
    highway_from = st.sidebar.selectbox("出発地点", highway_sections, index=oita_index)
    highway_to = st.sidebar.selectbox("到着地点", highway_sections, index=hita_index)
    
    st.sidebar.header("料金設定")
    # 片道料金設定
    one_way_fee = st.sidebar.number_input("片道料金（円）", min_value=0, value=2680, step=10)
    
    # メインエリア
    col1, col2 = st.columns([2, 1])
    
    with col1:
        st.header("CSVファイルのアップロード")
        uploaded_file = st.file_uploader("ETCカード利用明細CSVファイルをアップロードしてください", type=['csv'])
        
        if uploaded_file is not None:
            # エンコーディング検出
            encoding = detect_encoding(uploaded_file)
            st.info(f"検出されたエンコーディング: {encoding}")
            
            # CSVデータ読み込み
            df = load_csv_data(uploaded_file, encoding)
            
            if df is not None:
                # 年月を抽出
                year, month = extract_year_month(df)
                
                if year and month:
                    st.success(f"データ期間: {year}年{month}月")
                    
                    # データプレビュー
                    st.subheader("データプレビュー")
                    st.dataframe(df.head(10))
                    
                    # 統計情報
                    total_records = len(df)
                    total_fee = df['通行料金'].sum()
                    
                    col1_stat, col2_stat, col3_stat = st.columns(3)
                    with col1_stat:
                        st.metric("総利用回数", f"{total_records}回")
                    with col2_stat:
                        st.metric("総利用料金", f"¥{total_fee:,}")
                    with col3_stat:
                        expected_trips = 112560 // one_way_fee  # 固定値を使用
                        st.metric("想定利用回数", f"{expected_trips}回")
                    
                    # 実績簿生成ボタン
                    if st.button("利用実績簿を生成", type="primary"):
                        try:
                            wb = generate_expense_report_from_template(df, year, month, highway_from, highway_to, one_way_fee, 112560, organization, position, name)
                            
                            # Excelファイルをバイナリデータに変換
                            excel_buffer = io.BytesIO()
                            wb.save(excel_buffer)
                            excel_buffer.seek(0)
                            
                            # ダウンロードボタン
                            st.download_button(
                                label="📥 Excelファイルをダウンロード",
                                data=excel_buffer.getvalue(),
                                file_name=f"高速道路利用実績簿_{year}年{month}月.xlsx",
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                            )
                            
                            st.success("利用実績簿が正常に生成されました！")
                            
                        except Exception as e:
                            st.error(f"実績簿の生成に失敗しました: {e}")
                else:
                    st.error("データから年月を抽出できませんでした。")
    
    with col2:
        st.header("使用方法")
        st.markdown("""
        1. **CSVファイルをアップロード**
           - ETCカード利用明細のCSVファイルを選択
        
        2. **設定を確認**
           - 出発地点・到着地点を選択
           - 片道料金を入力
           - 月間認定額を入力
        
        3. **実績簿を生成**
           - 「利用実績簿を生成」ボタンをクリック
           - Excelファイルをダウンロード
        """)
        
        st.markdown("---")
        st.subheader("現在の設定")
        if organization:
            st.write(f"**所属:** {organization}")
        if position:
            st.write(f"**職:** {position}")
        if name:
            st.write(f"**氏名:** {name}")
        st.write(f"**利用区間:** {highway_from} ⇔ {highway_to}")
        st.write(f"**片道料金:** ¥{one_way_fee:,}")
        st.write(f"**月間認定額:** ¥112,560")
        
        st.markdown("---")
        st.subheader("📋 テンプレート準拠")
        st.markdown("""
        **正式テンプレート使用**
        - 公式フォーマット完全準拠
        - 水色箇所のみデータ入力
        - 数式・レイアウト保持
        - Excel機能完全再現
        """)

if __name__ == "__main__":
    main()